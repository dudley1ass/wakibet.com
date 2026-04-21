import csv
import json
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Set, Tuple

from openpyxl import load_workbook


def norm(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def slug(text: str) -> str:
    return "".join(ch for ch in text.upper() if ch.isalnum())[:8] or "DIV"


def main() -> None:
    source_path = Path(r"C:\Users\robin\Downloads\Winter_Springs_Player_Data_Improved.xlsx")
    out_dir = Path(r"d:\Winton Engineering\pickelball power\wakibet\apps\api\data")
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(source_path, data_only=True)
    ws = wb["Registrations"]

    division_players: Dict[Tuple[str, str, str, str], Set[str]] = defaultdict(set)
    players_seen = set()

    # Row 3 is the header row for "Registrations".
    for row in ws.iter_rows(min_row=4, values_only=True):
        player_name = norm(row[0])
        event_type = norm(row[5])
        skill_level = norm(row[6])
        age_bracket = norm(row[7])
        event_date = norm(row[8])

        if not player_name or not event_type:
            continue

        division = (event_type, skill_level, age_bracket, event_date)
        division_players[division].add(player_name)
        players_seen.add(player_name)

    matches: List[dict] = []
    per_player_matches: Dict[str, List[dict]] = defaultdict(list)

    for division, players_set in division_players.items():
        players = sorted(players_set)
        if len(players) < 2:
            continue

        event_type, skill_level, age_bracket, event_date = division
        base = f"{slug(event_type)}-{slug(event_date)}"
        match_num = 1

        # Test-run logic: round-robin by individual players within each division.
        for i in range(len(players)):
            for j in range(i + 1, len(players)):
                p1 = players[i]
                p2 = players[j]
                match_id = f"{base}-{match_num}"
                match_num += 1

                matches.append(
                    {
                        "match_id": match_id,
                        "event_type": event_type,
                        "skill_level": skill_level,
                        "age_bracket": age_bracket,
                        "event_date": event_date,
                        "player_a": p1,
                        "player_b": p2,
                        "status": "scheduled",
                    }
                )

                per_player_matches[p1].append(
                    {
                        "match_id": match_id,
                        "event_type": event_type,
                        "skill_level": skill_level,
                        "age_bracket": age_bracket,
                        "event_date": event_date,
                        "opponent": p2,
                    }
                )
                per_player_matches[p2].append(
                    {
                        "match_id": match_id,
                        "event_type": event_type,
                        "skill_level": skill_level,
                        "age_bracket": age_bracket,
                        "event_date": event_date,
                        "opponent": p1,
                    }
                )

    player_rows = []
    for player, player_matches in sorted(per_player_matches.items()):
        for m in player_matches:
            player_rows.append(
                {
                    "player_name": player,
                    "match_id": m["match_id"],
                    "event_type": m["event_type"],
                    "skill_level": m["skill_level"],
                    "age_bracket": m["age_bracket"],
                    "event_date": m["event_date"],
                    "opponent": m["opponent"],
                }
            )

    summary = {
        "tournament_name": "Winter Springs Spring Classic (test run)",
        "source_file": str(source_path),
        "players_in_registrations": len(players_seen),
        "players_with_generated_matches": len(per_player_matches),
        "divisions_considered": len(division_players),
        "matches_generated": len(matches),
        "generation_mode": "round_robin_by_player_within_division",
    }

    json_path = out_dir / "winter_springs_test_run_matches.json"
    csv_path = out_dir / "winter_springs_player_match_list.csv"
    summary_path = out_dir / "winter_springs_test_run_summary.json"

    json_path.write_text(
        json.dumps(
            {
                "summary": summary,
                "matches": matches,
                "per_player_matches": dict(sorted(per_player_matches.items())),
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "player_name",
                "match_id",
                "event_type",
                "skill_level",
                "age_bracket",
                "event_date",
                "opponent",
            ],
        )
        writer.writeheader()
        writer.writerows(player_rows)

    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    print(f"Wrote: {json_path}")
    print(f"Wrote: {csv_path}")
    print(f"Wrote: {summary_path}")


if __name__ == "__main__":
    main()
